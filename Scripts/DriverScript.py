import webbrowser
from threading import Thread
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
import datetime
import time
import os,sys
import XlsxReader
from XlsxReader import *
import Keywords
from Keywords import *
from openpyxl import load_workbook
import mimetypes,smtplib
import email
from email.mime.text import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email import Encoders
from email.MIMEBase import MIMEBase
import logging
import logging.handlers
import Config
from Config import *
import Constants
from Constants import *
from BackEndDrivers import *
#import Sikuli

logging.basicConfig(filename=LOG_FILE,level=logging.INFO,format='%(asctime)s %(levelname)s %(message)s',filemode='w')
logger=logging.getLogger(__name__)
handler=logging.handlers.RotatingFileHandler(LOG_FILE,backupCount=1)
logger.addHandler(handler)

StartTime=datetime.datetime.now()

def executefunctions(browser, driver):
	logger.info("Executing Driver Script: browser Name= "+browser)
	AttachmentsDirPath=DriverScript(browser, driver)
	logger.info("Test Case Execution Completed. Sending Email To Stake Holder After Test Case Execution: AttachmentsDirPath= "+AttachmentsDirPath)
	designgraphs(AttachmentsDirPath)
	#send_mail(AttachmentsDirPath,subject="Cureatr "+ browser +" Browser || Automaton Test Report")
	logger.info("Send Email Completed")

def DriverScript(browser, driver):
	#GET CURRENT DATE AND TIME
	currentdate=StartTime.strftime('%d-%m-%Y')
	currenttime=str(StartTime.time()).replace(":","")

	logger.info("CREATING FOLDER IN:" +OutPutFileDir+ " WITH TODAYS DATE")
	directory=OutPutFileDir+unicode(currentdate)+"/"
	#CREATE OUTPUT FOLDER WITH CURRENT TIME STAMP
	try:
		if not os.path.exists(directory):
			os.makedirs(directory)
			logger.info("Created Directory With Current Date: "+directory)
		subdirectory=directory+browser+"-Results-"+unicode(currenttime)[:6]+"/"
		if not os.path.exists(subdirectory):
			os.makedirs(subdirectory)
		logger.info("Created Sub Directory With Current Time For "+browser+ " Browser: "+subdirectory)
		logger.info("Copied Suite Xlsx and Results From Input Folder to " +browser+ " Output Folder")
	except:
		logger.info("Error: unable to create test results folder for:"+browser)
	
	copyWorkBook(Suite_Web, subdirectory)
	copyWorkBook(Results, subdirectory)
	if browser=="FF":
		time.sleep(0.3)
	elif browser=="IE":
		time.sleep(0.1)
		
	#READ SUITE XLS FILE FROM OUTPUT DIRECTORY
	SuiteXLSPath = subdirectory+'Suite_Web.xlsx'
	logger.info("Reading Suite sheet from Suite.xlsx workbook")
	SuiteXLS = load_workbook(SuiteXLSPath).get_sheet_by_name('Suite')
	#FOR LOOP TO GET/READ TEST SCRIPT NAME/ID FROM SUITE.XLS SHEET
	driver1="Test"
	driver2="Test"
	logger.info("Looping Suite Sheet to Get Test Script ID's Which Has Run Mode Yes")
	for num in range(2, SuiteXLS.max_row+1):
			currentTestSuite=getCellValue(SuiteXLSPath, "Suite", num, "TSID")
			if getCellValue(SuiteXLSPath, "Suite", num, "Runmode")=="Y":
				logger.info("Copying "+currentTestSuite+".xlsx sheet to "+browser+" output Folder")
				src=InPutFileDir+currentTestSuite+".xlsx"
				copyWorkBook(src, subdirectory)
				logger.info("Reading "+currentTestSuite+".xlsx sheet from "+browser+" output Folder")
				currentTestSuiteXLSPATH=subdirectory+currentTestSuite+".xlsx"
				currentTestSuiteXLS=load_workbook(currentTestSuiteXLSPATH).get_sheet_by_name('TestCases')
				logger.info("Reading TestCases sheet from " +currentTestSuite+ ".xlsx workbook")
				for count in range(2, currentTestSuiteXLS.max_row+1):
					currentTestCase=getCellValueBySheet(currentTestSuiteXLS, count, "TCID")
					Priority=getCellValueBySheet(currentTestSuiteXLS, count, "Priority")
					logger.info("Checking TestCase Runmode from " +currentTestSuite+ ".xlsx workbook")
					if getCellValueBySheet(currentTestSuiteXLS, count, "Runmode")=="Y":
						logger.info("Checking Test Data Sheet Exists For Current TestCase from " +currentTestSuite+ ".xlsx workbook")
						if issheetExist(currentTestSuiteXLSPATH, currentTestCase)==True:
							currentTestDataSheet=load_workbook(currentTestSuiteXLSPATH).get_sheet_by_name(currentTestCase)
							logger.info("Looping " +str(currentTestDataSheet) + " Based on Row Count")
							for dataset in range(2, currentTestDataSheet.max_row+1):
								logger.info("Checking Test Data Runmode from " + str(currentTestDataSheet) + " sheet")
								if getCellValueBySheet(currentTestDataSheet, dataset, "Runmode")=="Y":
									Correct_Data=getCellValueBySheet(currentTestDataSheet, dataset, "Correct_Data")
									DSID=getCellValueBySheet(currentTestDataSheet, dataset, "DSID")
									logger.info("Calling executeKeywords function")
									driverval=executeKeywords(currentTestSuiteXLSPATH, currentTestCase, browser, driver, dataset, count, Priority, subdirectory, currentTestDataSheet, DSID, driver1, driver2)
									driver1=driverval[0]
									driver2=driverval[1]
								else:
									logger.info("Printing NO RUN Status For "+ DSID +" Data Set Runmode No")
									currentTestSuiteSteps=load_workbook(currentTestSuiteXLSPATH).get_sheet_by_name('TestSteps')
									for TestStepsCount in range(2, currentTestSuiteSteps.max_row+1):
										if getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "TCID")==currentTestCase:
											resultSet=["NO RUN"]
											logger.info("Calling PrintTestStepResultNoRun Function To Print NORUN Status in Test Steps Sheet")
											#PrintTestStepResultNoRun(currentTestSuiteXLSPATH, currentTestSuiteSteps, Result, currentTestCase, dataset, TestStepsCount)
											PrintTestStepResult(currentTestSuiteXLSPATH, currentTestSuiteSteps, resultSet, currentTestCase, dataset, TestStepsCount)
											break
									
									columnname="Result"+str(dataset-1)
									addCellValue(currentTestSuiteXLSPATH, "TestCases", count, columnname, "NO RUN")
									addCellValue(currentTestSuiteXLSPATH, currentTestCase, dataset, "Result", "NO RUN")
									ResultsXLSPATH=subdirectory+'Results.xlsx'
									ResultsSheetXLS=load_workbook(ResultsXLSPATH).get_sheet_by_name('Status')
									NORUN=getCellValueBySheet(ResultsSheetXLS, 4, "NORUN")
									addCellValue(ResultsXLSPATH, "Status", 4, "NORUN", int(NORUN)+1)
						else:
							logger.info("Calling executeKeywords function when Data sheet not Exists")
							dataset=2
							currentTestDataSheet="Not Exists"
							DSID=""
							driverval=executeKeywords(currentTestSuiteXLSPATH, currentTestCase, browser, driver, dataset, count, Priority, subdirectory, currentTestDataSheet, DSID, driver1, driver2)
							driver1=driverval[0]
							driver2=driverval[1]
    						
	if driver1!="Test":
		logger.info("Closing dirver after test cases execution: "+browser)
		CloseBrowser(driver1)
	if driver2!="Test":
		logger.info("Closing dirver after test cases execution: "+browser)
		CloseBrowser(driver2)
	return subdirectory
    
def executeKeywords(currentTestSuiteXLSPATH, currentTestCase, browser, driver, dataset, count, Priority, subdirectory, currentTestDataSheet, DSID, driver1, driver2):
	currentTestSuiteSteps=load_workbook(currentTestSuiteXLSPATH).get_sheet_by_name('TestSteps')
	resultSet=[]
	logger.info("Looping "+currentTestCase+" Steps")
	for TestStepsCount in range(2, currentTestSuiteSteps.max_row+1):
		if getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "TCID")==currentTestCase:
			Keywords=getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "Keyword")
			target=getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "Target")
			data=getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "Data")
			user=getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "browser")
			TCID=getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "TCID")
			TSID=getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "TSID")
			Proceed_ON_FAIL=getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "Proceed_ON_FAIL")
			if currentTestDataSheet!="Not Exists":
				Correct_Data=getCellValueBySheet(currentTestDataSheet, dataset, "Correct_Data")
			else:
				Correct_Data="Y"
			if data is not None:
				if str(data).startswith("col"):
					datacolname=data.split("$")[1:][0]
					data=getCellValueBySheet(currentTestDataSheet, dataset, datacolname)
			if data is None:
				data=""
			logger.info("Current Test Step Values:: TCID: "+str(TCID)+" ,Keyword: "+str(Keywords)+" ,Target:"+str(target)+" ,Data :"+str(data)+" ,browser: "+str(browser)+" ,TSID: "+str(TSID)+" ,Proceed_ON_FAIL: "+str(Proceed_ON_FAIL)+" ,Correct_Data: "+str(Correct_Data))
			possibles = globals().copy()
			possibles.update(locals())
			if Keywords=="LaunchWebBrowser":
				method = possibles.get(Keywords)
				if driver1=="Test" and user=="user1":
					logger.info("Launching "+browser+ " Browser")
					retun=method(browser, driver, target, data, subdirectory, TCID, TSID, DSID, Correct_Data, currentTestDataSheet, dataset, user)
					driver1=retun[0]
					Keyword_execution_result_main=retun[1]
					resultSet.append(Keyword_execution_result_main)
				elif driver2=="Test" and user=='user2':
					logger.info("Launching 2nd "+browser+ " Browser")
					retun=method(browser, driver, target, data, subdirectory, TCID, TSID, DSID, Correct_Data, currentTestDataSheet, dataset, user)
					driver2=retun[0]
					Keyword_execution_result_main=retun[1]
					resultSet.append(Keyword_execution_result_main)
				else:
					logger.info(browser+ " Browser Already Running")
					Keyword_execution_result_main="NO RUN"
					resultSet.append(Keyword_execution_result_main)
					
				if (Proceed_ON_FAIL=="NO" and Keyword_execution_result_main=="FAIL"):
					logger.info("Stopping Current Test Cases Execution because Test Step is Failed and Proceed_ON_FAIL=NO")
					TestStepsCount=currentTestSuiteSteps.max_row+2
					method = possibles.get("CloseWebApp")
					retun=method(browser, driver, target, data, subdirectory, TCID, TSID, DSID, Correct_Data, currentTestDataSheet, dataset, user)
					break
			
			elif Keywords=="CreateUserPY" or Keywords=="CreateInstitution":
				method = possibles.get(Keywords)
				retun=method(browser, target, data, currentTestDataSheet, dataset,currentTestSuiteXLSPATH,currentTestCase)
				Keyword_execution_result_main=retun[0]
				resultSet.append(Keyword_execution_result_main)
				if retun[0] == "FAIL":
					send_mail(subdirectory,subject=Keywords+" Failed and stopped test exceution")					
					sys.exit(1)

			else:
				if user=="user1":
					driver=driver1
				elif user=="user2":
					driver=driver2
				method = possibles.get(Keywords)
				retun=method(browser, driver, target, data, subdirectory, TCID, TSID, DSID, Correct_Data, currentTestDataSheet, dataset, user)
				Keyword_execution_result_main=retun[0]
				Proceed_Next_Step=retun[1]
				resultSet.append(Keyword_execution_result_main)
				if (Proceed_ON_FAIL=="NO" and Keyword_execution_result_main=="FAIL") or Proceed_Next_Step=="NO":
					logger.info("Stopping Current Test Cases Execution because Test Step is Failed and Proceed_ON_FAIL=NO")
					TestStepsCount=currentTestSuiteSteps.max_row+2
					method = possibles.get("CloseWebApp")
					retun=method(browser, driver, target, data, subdirectory, TCID, TSID, DSID, Correct_Data, currentTestDataSheet, dataset, user)
					break

	logger.info("Calling PrintTestStepResult Function To Print PASS/FAIL Status in Test Steps Sheet")
	PrintTestStepResult(currentTestSuiteXLSPATH, currentTestSuiteSteps, resultSet, currentTestCase, dataset, TestStepsCount)
	logger.info("Calling printresults Function To Print PASS/FAIL Status In Results Sheet")
	printresults(currentTestSuiteXLSPATH, dataset, count, currentTestCase, resultSet, Priority, subdirectory, currentTestDataSheet,browser)
	return driver1, driver2

def PrintTestStepResult(currentTestSuiteXLSPATH, currentTestSuiteSteps, resultSet, currentTestCase, dataset, TestStepsCount):
	columnname="Result"+str(dataset-1)
	addColumn(currentTestSuiteXLSPATH, "TestSteps", columnname)
	addColumn(currentTestSuiteXLSPATH, "TestCases", columnname)
	addCellValueSet(currentTestSuiteXLSPATH, "TestSteps", TestStepsCount, columnname, resultSet, currentTestSuiteSteps, currentTestCase)
	
def printresults(currentTestSuiteXLSPATH, dataset, count, currentTestCase, resultSet, Priority, subdirectory, currentTestDataSheet,browser):
	ResultsXLSPATH=subdirectory+'Results.xlsx'
	ResultsSheetXLS=load_workbook(ResultsXLSPATH).get_sheet_by_name('Status')
	columnname="Result"+str(dataset-1)
	if "FAIL" in str(resultSet):
		addCellValue(currentTestSuiteXLSPATH, "TestCases", count, columnname, "FAIL")
		if currentTestDataSheet!="Not Exists":
			addCellValue(currentTestSuiteXLSPATH, currentTestCase, dataset, "Result", "FAIL")

		FAIL=getCellValueBySheet(ResultsSheetXLS, 4, "FAIL")
		addCellValue(ResultsXLSPATH, "Status", 4, "FAIL", int(FAIL)+1)
		
		P1=getCellValueBySheet(ResultsSheetXLS, 4, "P1")
		P2=getCellValueBySheet(ResultsSheetXLS, 4, "P2")
		P3=getCellValueBySheet(ResultsSheetXLS, 4, "P3")
		P4=getCellValueBySheet(ResultsSheetXLS, 4, "P4")
		if Priority=="P1":
			addCellValue(ResultsXLSPATH, "Status", 4, "P1", int(P1)+1)
		elif Priority=="P2":
			addCellValue(ResultsXLSPATH, "Status", 4, "P2", int(P2)+1)
		elif Priority=="P3":
			addCellValue(ResultsXLSPATH, "Status", 4, "P3", int(P3)+1)
		elif Priority=="P4":
			addCellValue(ResultsXLSPATH, "Status", 4, "P4", int(P4)+1)
	else:
		addCellValue(currentTestSuiteXLSPATH, "TestCases", count, columnname, "PASS")
		if currentTestDataSheet!="Not Exists":
			addCellValue(currentTestSuiteXLSPATH, currentTestCase, dataset, "Result", "PASS")

		PASS=getCellValueBySheet(ResultsSheetXLS, 4, "PASS")
		addCellValue(ResultsXLSPATH, "Status", 4, "PASS", int(PASS)+1)
		
	global StartTime
	currentdate=StartTime.strftime('%d-%b-%Y')
	EndTime=datetime.datetime.now()
	addCellValue(ResultsXLSPATH, "Status", 28, "Report", currentdate)
	addCellValue(ResultsXLSPATH, "Status", 29, "Report", StartTime.strftime('%X'))
	addCellValue(ResultsXLSPATH, "Status", 30, "Report", EndTime.strftime('%X'))
	elapsed=EndTime-StartTime
	addCellValue(ResultsXLSPATH, "Status", 31, "Report", elapsed)
	addCellValue(ResultsXLSPATH, "Status", 32, "Report", browser)
	
def send_mail(AttachmentsDirPath,subject):
    global logger
    logger.info('Gathering file and creating mail')
    msg=MIMEMultipart()
    msg['Subject']=subject
    msg['From']='cureatrtest@gmail.com'
    recipients=Constants.EmailAccountsForReports
    msg['To']=", ".join(recipients)
    msg.attach(MIMEText( 'PFA..Test Reports.This mail generated by Automation scripts'))
    for file in os.listdir(AttachmentsDirPath):
        filepath=os.path.join(AttachmentsDirPath,file)
        if not os.path.isfile(filepath):
        	continue
       	msgpart=MIMEBase('application',"vnd.openxmlformats-officedocument.spreadsheetml.sheet")
       	msgpart.set_payload(open(filepath,'rb').read())
       	Encoders.encode_base64(msgpart)
       	msgpart.add_header('Content-Disposition','attachment',filename=file)
       	msg.attach(msgpart)
    logger.info('Connecting to SMTP  server')
    server=smtplib.SMTP('smtp.gmail.com:587')
    server.starttls()
    server.login('cureatrtest@gmail.com','Cureatr@1234')
    server.sendmail(msg['From'],recipients,msg.as_string())	
    logger.info('mail sent with attachments')
    server.quit()
		   		
# open a public URL, in this case, the webbrowser docs

if __name__ == '__main__':
	try:
		t1=Thread(target=executefunctions,args=('Chrome', ''))
		#t2=Thread(target=executefunctions,args=('FF', ''))
		#t3=Thread(target=executefunctions,args=('IE', ''))
		t1.start()
		#t2.start()
		#t3.start()
	except:
		print "Error: unable to start thread"
