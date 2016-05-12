#! /usr/bin/python
import os,sys
import shutil
import openpyxl
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import  coordinate_from_string,column_index_from_string,get_column_letter
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.layout import Layout, ManualLayout

def copyWorkBook(src, dst):
    shutil.copy(src, dst)

def getRowCount(sh):
    row_count= sh.max_row
    print row_count

def getColCount(sh):
    col_count=sh.max_column
    print col_count

def getCellData():
    for row in sh.iter_rows() :
        for cell in row:
            print cell.value
            
def Xlsx_Reader(FilePath, SheetName):
	wb=load_workbook(FilePath)
	sh=wb[SheetName]
	return sh
	
def issheetExist(FilePath, SheetName):
	 wb1 = load_workbook(FilePath)
	 listSheet=wb1.get_sheet_names()
	 for sname in range(0,len(listSheet)):
	 	if SheetName == listSheet[sname]:
	 		sheetExist="Yes"
	 		break
	 	else:
	 		sheetExist="No"
	 		
	 	
	 if sheetExist=="Yes":
	 	return True
	 else:
	 	return False

def getCellValue(wb1, sh1, row, column):
     val=[]
     wb = load_workbook(wb1)
     sh2= wb[sh1]
     for srow in range(1, sh2.max_column+1):
         #for cell in srow:
             if sh2.cell(row=1,column=srow).value == column:
               val=sh2.cell(row=1, column=srow).col_idx
     return sh2.cell(row=row, column=val).value

def getCellValueBySheet(sh1, row, column):
	val=[]
	for srow in range(1, sh1.max_column+1):
		#for cell in srow:
		if sh1.cell(row=1,column=srow).value == column:
			val=sh1.cell(row=1, column=srow).col_idx

	return sh1.cell(row=row, column=val).value

def addColumn(wbpath, sh, column):
	wb=load_workbook(wbpath)
	sh=wb.get_sheet_by_name(sh)
	cl=sh.max_column
	for srow in range(1, sh.max_column+1):
         #for cell in srow:
         	if sh.cell(row=1,column=srow).value == column:
         		val=sh.cell(row=1, column=srow).col_idx
         		createsheet="No"
         		break
         	else:
         		createsheet="Yes"
         	
	if createsheet=="Yes":
		sh.cell(row=1,column=cl+1).value=column
		wb.save(wbpath)

def addCellValue(wbpath, sheet, row, column, columnvalue):
	val=[]
	wb=load_workbook(wbpath)
	sh=wb.get_sheet_by_name(sheet)
	for srow in range(1, sh.max_column+1):
		#for cell in srow:
		if sh.cell(row=1,column=srow).value == column:
			val=sh.cell(row=1, column=srow).col_idx
			sh.cell(row=row,column=val).value=columnvalue
		
	wb.save(wbpath)

def addCellValueToBuff(sh, row, column, columnvalue):
	val=[]
	for srow in range(1, sh.max_column+1):
		#for cell in srow:
		if sh.cell(row=1,column=srow).value == column:
			val=sh.cell(row=1, column=srow).col_idx
			sh.cell(row=row,column=val).value=columnvalue
		
             
def addCellValueSet(wbpath, sheet, row, column, resultSet, currentTestSuiteSteps, currentTestCase):
	val=[]
	wb=load_workbook(wbpath)
	sh=wb.get_sheet_by_name(sheet)
	resultid=0
	for row in range(2, currentTestSuiteSteps.max_row+1):
		if getCellValueBySheet(currentTestSuiteSteps, row, "TCID")==currentTestCase:
			for srow in range(1, sh.max_column+1):
				#for cell in srow:
				if sh.cell(row=1,column=srow).value == column:
					val=sh.cell(row=1, column=srow).col_idx
					sh.cell(row=row,column=val).value=resultSet[resultid]
					
			resultid=resultid+1
			if len(resultSet)==resultid:
				row=currentTestSuiteSteps.max_row+10
				break
				
	wb.save(wbpath)
	
def addCellValueNoRun(wbpath, sheet, row, column, Result, currentTestSuiteSteps, currentTestCase):
	val=[]
	wb=load_workbook(wbpath)
	sh=wb.get_sheet_by_name(sheet)
	for row in range(2, currentTestSuiteSteps.max_row+1):
		if getCellValueBySheet(currentTestSuiteSteps, row, "TCID")==currentTestCase:
			for srow in range(1, sh.max_column+1):
				#for cell in srow:
				if sh.cell(row=1,column=srow).value == column:
					val=sh.cell(row=1, column=srow).col_idx
					sh.cell(row=row,column=val).value=Result
					
	wb.save(wbpath)
	
def designgraphs(wbpath):
	wbpath=wbpath+'Results.xlsx'
	wb = load_workbook(wbpath)
	ws=wb.active
	StatusData = Reference(ws, min_col=1, min_row=3, max_col=3, max_row=4)
	StatusChart = BarChart()
	StatusChart.style = 10
	StatusChart.title = "Test Execution Status"
	StatusChart.y_axis.title = 'Test cases executed'
	StatusChart.x_axis.title = 'Test case status'
	StatusChart.add_data(StatusData,titles_from_data=True)
	ws.add_chart(StatusChart,"A7")
	
	BugsData = Reference(ws, min_col=7, min_row=3, max_col=10, max_row=4)
	BugsPriorityChart = BarChart()
	BugsPriorityChart.style = 10
	BugsPriorityChart.title = "Test Execution Status"
	BugsPriorityChart.y_axis.title = 'Test cases executed'
	BugsPriorityChart.x_axis.title = 'Test case status'
	BugsPriorityChart.add_data(BugsData,titles_from_data=True)
	ws.add_chart(BugsPriorityChart,"F7")
	wb.save(wbpath)