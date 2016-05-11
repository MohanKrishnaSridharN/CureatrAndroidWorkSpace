import webbrowser
from threading import Thread
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
import datetime
import time
import os,sys
import XlsxReader
from XlsxReader import *
import Keywords
from Keywords import *
import createuser
from createuser import *
from openpyxl import load_workbook
import mimetypes,smtplib
import email
from email.mime.text import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email import Encoders
from email.MIMEBase import MIMEBase
import logging
import logging.handlers
import Config
from Config import *
sys.path.insert(0,'/Users/macmini/dev/cureatr/server/qa/')
import db_recipes
	
#logdir="/Users/macmini/Cureatr/CureatrPythonWorkSpace/applicationlogs"
LOG_FILE=getattr(Config, str("LOG_FILE"))
logging.basicConfig(level=logging.INFO,filename=LOG_FILE)
logger=logging.getLogger(__name__)
#handler=logging.handlers.RotatingFileHandler(LOG_File,maxBytes=20,backupCount=5)
#logger.addHandler(handler)

mylist = []
StartTime=datetime.datetime.now()

def executefunctions(browser, driver):
	logger.info("Execute Driver Script::"+"browser Name="+browser+"Driver Name="+driver)
	AttachmentsDirPath=DriverScript(browser, driver)
	logger.info("Test Case Executiong Compleated::"+"Sending Email To Stake Holder After Test Case Execution::"+"AttachmentsDirPath="+AttachmentsDirPath)
	#designgraphs(AttachmentsDirPath)
	#send_mail(AttachmentsDirPath)
	logger.info("Send Email Compleated")

def DriverScript(browser, driver):
	#GET CURRENT DATE AND TIME
	currenttime1=datetime.datetime.now().time()
	currentdate=datetime.date.today()
	currenttime=str(currenttime1).replace(":","")
	#startTime=time.time()
	
	#CREATE SUBFOLDER IN OUTPUTFILES FOLDER WITH TODAYS DATE
	directory=getattr(Config, str("OutPutFileDir"))+unicode(currentdate)+"/"
	try:
		if browser=="FF":
			if not os.path.exists(directory):
				os.makedirs(directory)
		elif browser=="Chrome":
			time.sleep(0.2)
			if not os.path.exists(directory):
				os.makedirs(directory)
		elif browser== "IE":
			time.sleep(0.1)
			if not os.path.exists(directory):
				os.makedirs(directory)
	except:
		print "Error: unable to create folder"
		
	#CREATE OUTPUT FOLDER WITH CURRENT TIME STAMP
	Suite_Web=getattr(Config, str("Suite_Web"))
	Results=getattr(Config, str("Results"))
	try:
		if browser=="FF":
			subdirectory=directory+"FF-Results-"+unicode(currenttime)[:6]+"/"
			os.makedirs(subdirectory)
			dst=subdirectory
			#COPYING SUITE XLSX FILE FROM INPUT FOLDER TO OUTPUT FOLDER
			copyWorkBook(Suite_Web, dst)
			#COPYING RESULTS XLSX FILE FROM INPUT FOLDER TO OUTPUT FOLDER
			copyWorkBook(Results, dst)
		elif browser=="Chrome":
			subdirectory=directory+"Chrome-Results-"+unicode(currenttime)[:6]+"/"
			os.makedirs(subdirectory)
			dst=subdirectory
			#COPYING SUITE XLSX FILE FROM INPUT FOLDER TO OUTPUT FOLDER
			copyWorkBook(Suite_Web, dst)
			#COPYING RESULTS XLSX FILE FROM INPUT FOLDER TO OUTPUT FOLDER
			copyWorkBook(Results, dst)
		elif browser=="IE":
			subdirectory=directory+"IE-Results-"+unicode(currenttime)[:6]+"/"
			os.makedirs(subdirectory)
			dst=subdirectory
			#COPYING SUITE XLSX FILE FROM INPUT FOLDER TO OUTPUT FOLDER
			copyWorkBook(Suite_Web, dst)
			#COPYING RESULTS XLSX FILE FROM INPUT FOLDER TO OUTPUT FOLDER
			copyWorkBook(Results, dst)
	except:
		print "Error: unable to create sub folder"
		
	#READ SUITE XLS FILE FROM OUTPUT DIRECTORY
	SuiteXLSPath = subdirectory+'Suite_Web.xlsx'
	#sh=load_workbook(FilePath).get_sheet_by_name('Suite')
	SuiteXLS = load_workbook(SuiteXLSPath).get_sheet_by_name('Suite')
	#FOR LOOP TO GET/READ TEST SCRIPT NAME/ID FROM SUITE.XLS SHEET
	driver1="Test"
	driver2="Test"
	for num in range(2, SuiteXLS.max_row+1):
			currentTestSuite=getCellValue(SuiteXLSPath, "Suite", num, "TSID")
			if getCellValue(SuiteXLSPath, "Suite", num, "Runmode")=="Y":
				#COPY CURRENTTESTSUITE XLS FILE FROM OUTPUT DIRECTORY
				src=getattr(Config, str("InPutFileDir"))+currentTestSuite+".xlsx"
				dst=subdirectory
				copyWorkBook(src, dst)
				#READ CURRENTTESTSUITE XLS FILE FROM OUTPUT DIRECTORY\
				currentTestSuiteXLSPATH=subdirectory+currentTestSuite+".xlsx"
				currentTestSuiteXLS=load_workbook(currentTestSuiteXLSPATH).get_sheet_by_name('TestCases')
				for count in range(2, currentTestSuiteXLS.max_row+1):
					#currentTestCase=currentTestSuiteXLS.cell(row=count, column=1)
					currentTestCase=getCellValueBySheet(currentTestSuiteXLS, count, "TCID")
					Priority=getCellValueBySheet(currentTestSuiteXLS, count, "Priority")
					if getCellValueBySheet(currentTestSuiteXLS, count, "Runmode")=="Y":
						if issheetExist(currentTestSuiteXLSPATH, currentTestCase)==True:
							currentTestDataSheet=load_workbook(currentTestSuiteXLSPATH).get_sheet_by_name(currentTestCase)
							for dataset in range(2, currentTestDataSheet.max_row+1):
								if getCellValueBySheet(currentTestDataSheet, dataset, "Runmode")=="Y":
									Correct_Data=getCellValueBySheet(currentTestDataSheet, dataset, "Correct_Data")
									#Createuser=getCellValueBySheet(currentTestDataSheet, dataset, "Createuser")
									DSID=getCellValueBySheet(currentTestDataSheet, dataset, "DSID")
									driverval=executeKeywords(currentTestSuiteXLSPATH, currentTestCase, browser, driver, dataset, count, Priority, subdirectory, currentTestDataSheet, DSID, driver1, driver2)
									driver1=driverval[0]
									driver2=driverval[1]
								else:
									currentTestSuiteSteps=load_workbook(currentTestSuiteXLSPATH).get_sheet_by_name('TestSteps')
									for TestStepsCount in range(2, currentTestSuiteSteps.max_row+1):
										if getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "TCID")==currentTestCase:
											#print "NO RUN"
											createXLSReport(currentTestSuiteXLSPATH, TestStepsCount, "NO RUN", dataset)
									
									#printresults(currentTestSuiteXLSPATH, dataset, count, currentTestCase, resultSet)
									columnname="Result"+str(dataset-1)
									addCellValue(currentTestSuiteXLSPATH, "TestCases", count, columnname, "NO RUN")
									addCellValue(currentTestSuiteXLSPATH, currentTestCase, dataset, "Result", "NO RUN")
									ResultsXLSPATH=subdirectory+'Results.xlsx'
									ResultsSheetXLS=load_workbook(ResultsXLSPATH).get_sheet_by_name('Status')
									NORUN=getCellValueBySheet(ResultsSheetXLS, 4, "NORUN")
									addCellValue(ResultsXLSPATH, "Status", 4, "NORUN", int(NORUN)+1)
						else:
							dataset=2
							currentTestDataSheet="Not Exists"
							DSID=""
							driverval=executeKeywords(currentTestSuiteXLSPATH, currentTestCase, browser, driver, dataset, count, Priority, subdirectory, currentTestDataSheet, DSID, driver1, driver2)
							driver1=driverval[0]
							driver2=driverval[1]
    						
	return subdirectory
    
def executeKeywords(currentTestSuiteXLSPATH, currentTestCase, browser, driver, dataset, count, Priority, subdirectory, currentTestDataSheet, DSID, driver1, driver2):
	currentTestSuiteSteps=load_workbook(currentTestSuiteXLSPATH).get_sheet_by_name('TestSteps')
	resultSet=[]
	for TestStepsCount in range(2, currentTestSuiteSteps.max_row+1):
		if getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "TCID")==currentTestCase:
			Keywords=getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "Keyword")
			target=getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "Target")
			data=getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "Data")
			user=getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "browser")
			TCID=getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "TCID")
			TSID=getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "TSID")
			Proceed_ON_FAIL=getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "Proceed_ON_FAIL")
			Correct_Data=getCellValueBySheet(currentTestDataSheet, dataset, "Correct_Data")
			if data is not None:
				if str(data).startswith("col"):
					datacolname=data.split("$")[1:][0]
					data=getCellValueBySheet(currentTestDataSheet, dataset, datacolname)
			
			possibles = globals().copy()
			possibles.update(locals())
			if Keywords=="LaunchWebBrowser":
				method = possibles.get(Keywords)
				if driver1=="Test" and user=="user1":
					retun=method(browser, driver, target, data, subdirectory, TCID, TSID, DSID, Correct_Data)
				
				if driver2=="Test" and user=='user2':
					retun=method(browser, driver, target, data, subdirectory, TCID, TSID, DSID, Correct_Data)
				
				if user=="user1" and driver1=="Test":
					driver1=retun[0]
					Keyword_execution_result_main=retun[1]
					resultSet.append(Keyword_execution_result_main)
					createXLSReport(currentTestSuiteXLSPATH, TestStepsCount, Keyword_execution_result_main, dataset)
				elif user=="user2" and driver2=="Test":
					driver2=retun[0]
					Keyword_execution_result_main=retun[1]
					resultSet.append(Keyword_execution_result_main)
					createXLSReport(currentTestSuiteXLSPATH, TestStepsCount, Keyword_execution_result_main, dataset)
					if Proceed_ON_FAIL=="NO" and Keyword_execution_result_main=="FAIL":
						TestStepsCount=currentTestSuiteSteps.max_row+2
						method = possibles.get("CloseWebApp")
						retun=method(browser, driver, target, data, subdirectory, TCID, TSID, DSID, Correct_Data)
						break

			elif Keywords=="CreateUserPY":
				data=getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "Data")
				if data is not None:
					if str(data).startswith("PY"):
						INSTITUTIONID=getCellValueBySheet(currentTestDataSheet, dataset, data.split("$")[1:][0])
						TYPE=getCellValueBySheet(currentTestDataSheet, dataset, data.split("$")[1:][1])
						TITILE=getCellValueBySheet(currentTestDataSheet, dataset, data.split("$")[1:][2])
						SPECIALTY=getCellValueBySheet(currentTestDataSheet, dataset, data.split("$")[1:][3])
						FIRSTNAME=getCellValueBySheet(currentTestDataSheet, dataset, data.split("$")[1:][4])
						LASTNAME=getCellValueBySheet(currentTestDataSheet, dataset, data.split("$")[1:][5])
						USERNAME=getCellValueBySheet(currentTestDataSheet, dataset, data.split("$")[1:][6])
						PASSWORD=getCellValueBySheet(currentTestDataSheet, dataset, data.split("$")[1:][7])
						EMAILID="test13985@mtuity.com"
						#db_recipes.qa_create_user()
						db_recipes.qa_create_user(first_name=FIRSTNAME, institution_id=INSTITUTIONID, specialty=SPECIALTY,
							admin_iids=None, title=TITILE, password=PASSWORD, device_token_type="tokens.TOKEN_TYPE_IOS",
							last_name=LASTNAME, email=EMAILID, uid=None, managed=False, pin=None, make_active=True,
							additional_device_token_user_ids=None, wowos=None, roles=None)

			elif Keywords=="CreateInstitution":
				data=getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "Data")
				if data is not None:
					if str(data).startswith("PY"):
						INSTITUTIONID=getCellValueBySheet(currentTestDataSheet, dataset, data.split("$")[1:][0])
						INSTITUTIONSHORTNAME=getCellValueBySheet(currentTestDataSheet, dataset, data.split("$")[1:][1])
						INSTITUTIONNAME=getCellValueBySheet(currentTestDataSheet, dataset, data.split("$")[1:][2])
						db_recipes.qa_create_institution(INSTITUTIONID, short_name=INSTITUTIONSHORTNAME, name=INSTITUTIONNAME, parent_iid="", master_iid="", tags=list())



			else:
				if user=="user1":
					driver=driver1
				elif user=="user2":
					driver=driver2
				method = possibles.get(Keywords)
				retun=method(browser, driver, target, data, subdirectory, TCID, TSID, DSID, Correct_Data)
				Keyword_execution_result_main=retun[0]
				Proceed_Next_Step=retun[1]
				resultSet.append(Keyword_execution_result_main)
				createXLSReport(currentTestSuiteXLSPATH, TestStepsCount, Keyword_execution_result_main, dataset)
				if (Proceed_ON_FAIL=="NO" and Keyword_execution_result_main=="FAIL") or Proceed_Next_Step=="NO":
					TestStepsCount=currentTestSuiteSteps.max_row+2
					method = possibles.get("CloseWebApp")
					retun=method(browser, driver, target, data, subdirectory, TCID, TSID, DSID, Correct_Data)
					break
				
	printresults(currentTestSuiteXLSPATH, dataset, count, currentTestCase, resultSet, Priority, subdirectory, currentTestDataSheet)
	return driver1, driver2

def createXLSReport(currentTestSuiteXLSPATH, TestStepsCount, Keyword_execution_result_main, dataset):
	columnname="Result"+str(dataset-1)
	addColumn(currentTestSuiteXLSPATH, "TestSteps", columnname)
	addColumn(currentTestSuiteXLSPATH, "TestCases", columnname)
	addCellValue(currentTestSuiteXLSPATH, "TestSteps", TestStepsCount, columnname, Keyword_execution_result_main)

def printresults(currentTestSuiteXLSPATH, dataset, count, currentTestCase, resultSet, Priority, subdirectory, currentTestDataSheet):
	ResultsXLSPATH=subdirectory+'Results.xlsx'
	ResultsSheetXLS=load_workbook(ResultsXLSPATH).get_sheet_by_name('Status')
	columnname="Result"+str(dataset-1)
	if "FAIL" in str(resultSet):
		addCellValue(currentTestSuiteXLSPATH, "TestCases", count, columnname, "FAIL")
		if currentTestDataSheet!="Not Exists":
			addCellValue(currentTestSuiteXLSPATH, currentTestCase, dataset, "Result", "FAIL")

		FAIL=getCellValueBySheet(ResultsSheetXLS, 4, "FAIL")
		addCellValue(ResultsXLSPATH, "Status", 4, "FAIL", int(FAIL)+1)
		
		P1=getCellValueBySheet(ResultsSheetXLS, 4, "P1")
		P2=getCellValueBySheet(ResultsSheetXLS, 4, "P2")
		P3=getCellValueBySheet(ResultsSheetXLS, 4, "P3")
		P4=getCellValueBySheet(ResultsSheetXLS, 4, "P4")
		if Priority=="P1":
			addCellValue(ResultsXLSPATH, "Status", 4, "P1", int(P1)+1)
		elif Priority=="P2":
			addCellValue(ResultsXLSPATH, "Status", 4, "P2", int(P2)+1)
		elif Priority=="P3":
			addCellValue(ResultsXLSPATH, "Status", 4, "P3", int(P3)+1)
		elif Priority=="P4":
			addCellValue(ResultsXLSPATH, "Status", 4, "P4", int(P4)+1)
	else:
		addCellValue(currentTestSuiteXLSPATH, "TestCases", count, columnname, "PASS")
		if currentTestDataSheet!="Not Exists":
			addCellValue(currentTestSuiteXLSPATH, currentTestCase, dataset, "Result", "PASS")

		PASS=getCellValueBySheet(ResultsSheetXLS, 4, "PASS")
		addCellValue(ResultsXLSPATH, "Status", 4, "PASS", int(PASS)+1)
		
	global StartTime
	currentdate=StartTime.strftime('%d-%b-%Y')
	EndTime=datetime.datetime.now()
	addCellValue(ResultsXLSPATH, "Status", 28, "Report", currentdate)
	addCellValue(ResultsXLSPATH, "Status", 29, "Report", StartTime.strftime('%X'))
	addCellValue(ResultsXLSPATH, "Status", 30, "Report", EndTime.strftime('%X'))
	elapsed=EndTime-StartTime
	addCellValue(ResultsXLSPATH, "Status", 31, "Report", elapsed)
	
def send_mail(AttachmentsDirPath):
    global logger
    logger.info('Gathering file and creating mail')
    msg=MIMEMultipart()
    msg['Subject']='Tesf file attached'
    msg['From']='techops@mtuity.com'
    recipients=['mohan.nimmala@mtuity.com', 'rithesh.karra@mtuity.com']
    msg['To']=", ".join(recipients)
    msg.attach(MIMEText( "Test mail"))
    for file in os.listdir(AttachmentsDirPath):
        filepath=os.path.join(AttachmentsDirPath,file)
        if not os.path.isfile(filepath):
           continue
        msgpart=MIMEBase('application',"vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        msgpart.set_payload(open(filepath,'rb').read())
        Encoders.encode_base64(msgpart)
        msgpart.add_header('Content-Disposition','attachment',filename=file)
        msg.attach(msgpart)
    logger.info('Connecting to SMTP  server')
    server=smtplib.SMTP('smtp.gmail.com:587')
    server.starttls()
    server.login('techops@mtuity.com','Paradigm@123.')
    server.sendmail(msg['From'],recipients,msg.as_string())
    logger.info('mail sent with attachments')
    server.quit()
		   		
# open a public URL, in this case, the webbrowser docs
if __name__ == '__main__':
	try:
		t1=Thread(target=executefunctions,args=('FF', ''))
		#t2=Thread(target=executefunctions,args=('Chrome', ''))
		#t3=Thread(target=executefunctions,args=('IE', ''))
		t1.start()
		#t2.start()
		#t3.start()
	except:
		print "Error: unable to start thread"
