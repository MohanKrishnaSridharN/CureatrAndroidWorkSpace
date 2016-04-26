import webbrowser
from threading import Thread
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
import datetime
import time
import os
#import ParallalBrowser
#from ParallalBrowser import *
import XlsxReader
from XlsxReader import *
import Keywords
from Keywords import *
from openpyxl import load_workbook
import mimetypes,smtplib
import email
from email.mime.text import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email import Encoders
from email.MIMEBase import MIMEBase
import logging
import logging.handlers



#logdir="/Users/macmini/Cureatr/CureatrPythonWorkSpace/applicationlogs"
LOG_FILE='/Users/macmini/Cureatr/CureatrPythonWorkSpace/applicationlogs/applog.txt'
logging.basicConfig(level=logging.INFO,filename=LOG_FILE)
logger=logging.getLogger(__name__)
#handler=logging.handlers.RotatingFileHandler(LOG_File,maxBytes=20,backupCount=5)
#logger.addHandler(handler)


mylist = []
StartTime=datetime.datetime.now()

def executefunctions(browser, driver):
    AttachmentsDirPath=DriverScript(browser, driver)
    send_mail(AttachmentsDirPath)
    
    
    #driver=openbrowser(browser, driver)
    #openapp(driver)
    #close(driver)

def DriverScript(browser, driver):
    #GET CURRENT DATE AND TIME 
    currenttime1=datetime.datetime.now().time()
    currentdate=datetime.date.today()
    currenttime=str(currenttime1).replace(":","")
    #startTime=time.time()
    
    #CREATE SUBFOLDER IN OUTPUTFILES FOLDER WITH TODAYS DATE
    directory="/Users/macmini/Cureatr/CureatrPythonWorkSpace/OutPutFiles/"+unicode(currentdate)+"/"
    try:
        if browser=="FF":
            if not os.path.exists(directory):
                os.makedirs(directory)
        elif browser=="Chrome":
            time.sleep(0.2)
            if not os.path.exists(directory):
                os.makedirs(directory)
        elif browser== "IE":
            time.sleep(0.1)
            if not os.path.exists(directory):
                os.makedirs(directory)
    except:
        print "Error: unable to create folder"

    #CREATE OUTPUT FOLDER WITH CURRENT TIME STAMP
    Suite_Web="/Users/macmini/Cureatr/CureatrPythonWorkSpace/InputFiles/Web/Suite_Web.xlsx"
    Results="/Users/macmini/Cureatr/CureatrPythonWorkSpace/InputFiles/Web/Results.xlsx"
    try:
        if browser=="FF":
            subdirectory=directory+"FF-Results-"+unicode(currenttime)[:6]+"/"
            os.makedirs(subdirectory)
            dst=subdirectory
            #COPYING SUITE XLSX FILE FROM INPUT FOLDER TO OUTPUT FOLDER
            copyWorkBook(Suite_Web, dst)
            #COPYING RESULTS XLSX FILE FROM INPUT FOLDER TO OUTPUT FOLDER
            copyWorkBook(Results, dst)
        elif browser=="Chrome":
            subdirectory=directory+"Chrome-Results-"+unicode(currenttime)[:6]+"/"
            os.makedirs(subdirectory)
            dst=subdirectory
            #COPYING SUITE XLSX FILE FROM INPUT FOLDER TO OUTPUT FOLDER
            copyWorkBook(Suite_Web, dst)
            #COPYING RESULTS XLSX FILE FROM INPUT FOLDER TO OUTPUT FOLDER
            copyWorkBook(Results, dst)
        elif browser=="IE":
            subdirectory=directory+"IE-Results-"+unicode(currenttime)[:6]+"/"
            os.makedirs(subdirectory)
            dst=subdirectory
            #COPYING SUITE XLSX FILE FROM INPUT FOLDER TO OUTPUT FOLDER
            copyWorkBook(Suite_Web, dst)
            #COPYING RESULTS XLSX FILE FROM INPUT FOLDER TO OUTPUT FOLDER
            copyWorkBook(Results, dst)
    except:
        print "Error: unable to create sub folder"

    #READ SUITE XLS FILE FROM OUTPUT DIRECTORY
    SuiteXLSPath = subdirectory+'Suite_Web.xlsx'
    #sh=load_workbook(FilePath).get_sheet_by_name('Suite')
    SuiteXLS = load_workbook(SuiteXLSPath).get_sheet_by_name('Suite')
        
    #FOR LOOP TO GET/READ TEST SCRIPT NAME/ID FROM SUITE.XLS SHEET
    for num in range(2, SuiteXLS.max_row+1):
           	currentTestSuite=getCellValue(SuiteXLSPath, "Suite", num, "TSID")
           	if getCellValue(SuiteXLSPath, "Suite", num, "Runmode")=="Y":
           		#COPY CURRENTTESTSUITE XLS FILE FROM OUTPUT DIRECTORY
           		src="/Users/macmini/Cureatr/CureatrPythonWorkSpace/InputFiles/Web/"+currentTestSuite+".xlsx"
           		dst=subdirectory
           		copyWorkBook(src, dst)
           		#READ CURRENTTESTSUITE XLS FILE FROM OUTPUT DIRECTORY
           		currentTestSuiteXLSPATH=subdirectory+currentTestSuite+".xlsx"
           		currentTestSuiteXLS=load_workbook(currentTestSuiteXLSPATH).get_sheet_by_name('TestCases')
           		for count in range(2, currentTestSuiteXLS.max_row+1):
           			#currentTestCase=currentTestSuiteXLS.cell(row=count, column=1)
           			currentTestCase=getCellValueBySheet(currentTestSuiteXLS, count, "TCID")
           			Priority=getCellValueBySheet(currentTestSuiteXLS, count, "Priority")
           			if getCellValueBySheet(currentTestSuiteXLS, count, "Runmode")=="Y":
           				if issheetExist(currentTestSuiteXLSPATH, currentTestCase)==True:
           					print "Test Data Sheet Exists"
           					currentTestDataSheet=load_workbook(currentTestSuiteXLSPATH).get_sheet_by_name(currentTestCase)
           					for dataset in range(2, currentTestDataSheet.max_row+1):
           						if getCellValueBySheet(currentTestDataSheet, dataset, "Runmode")=="Y":
           							Correct_Data=getCellValueBySheet(currentTestDataSheet, dataset, "Correct_Data")
           							#Createuser=getCellValueBySheet(currentTestDataSheet, dataset, "Createuser")
           							executeKeywords(currentTestSuiteXLSPATH, currentTestCase, browser, driver, dataset, count, Priority, subdirectory, currentTestDataSheet)
           						else:
           							currentTestSuiteSteps=load_workbook(currentTestSuiteXLSPATH).get_sheet_by_name('TestSteps')
           							for TestStepsCount in range(2, currentTestSuiteSteps.max_row+1):
           								if getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "TCID")==currentTestCase:
           									createXLSReport(currentTestSuiteXLSPATH, TestStepsCount, "NO RUN", dataset)
           									
           							#printresults(currentTestSuiteXLSPATH, dataset, count, currentTestCase, resultSet)
           							columnname="Result"+str(dataset-1)
           							addCellValue(currentTestSuiteXLSPATH, "TestCases", count, columnname, "NO RUN")
           							addCellValue(currentTestSuiteXLSPATH, currentTestCase, dataset, "Result", "NO RUN")
           							NORUN=getCellValueBySheet(ResultsSheetXLS, 7, "Result")
           							addCellValue(ResultsXLSPATH, "Status", 7, "Result", int(NORUN)+1)
           				else:
           					#This Condition Not Working
           					dataset=2
           					currentTestDataSheet=""
           					print "currentTestDataSheet Not exists"
           					executeKeywords(currentTestSuiteXLSPATH, currentTestCase, browser, driver, dataset, count, Priority, subdirectory, currentTestDataSheet)	
           					print "sheet not exists"
           					
    return subdirectory

def executeKeywords(currentTestSuiteXLSPATH, currentTestCase, browser, driver, dataset, count, Priority, subdirectory, currentTestDataSheet):
	currentTestSuiteSteps=load_workbook(currentTestSuiteXLSPATH).get_sheet_by_name('TestSteps')
	resultSet=[]
	driver1=[]
	for TestStepsCount in range(2, currentTestSuiteSteps.max_row+1):
		#resultSet.append("test")
		if getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "TCID")==currentTestCase:
			Keywords=getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "Keyword")
			target=getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "Target")
			data=getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "Data")
			user=getCellValueBySheet(currentTestSuiteSteps, TestStepsCount, "browser")
			if data is not None:
				if str(data).startswith("col"):
					datacolname=data.split("$")[1:][0]
					data=getCellValueBySheet(currentTestDataSheet, dataset, datacolname)
			
			possibles = globals().copy()
			possibles.update(locals())
			#driver2=''
			if Keywords=="LaunchWebBrowser":
				method = possibles.get(Keywords)
				retun=method(browser, driver, target, data)
				#user=retun[0]
				if user=="user1":
					driver1=retun[0]
				elif user=="user2":
					driver2=retun[0]
					Keyword_execution_result_main=retun[1]
					resultSet.append(Keyword_execution_result_main)
					createXLSReport(currentTestSuiteXLSPATH, TestStepsCount, Keyword_execution_result_main, dataset)
			else:
				if user=="user1":
					driver=driver1
				elif user=="user2":
					driver=driver2
				method = possibles.get(Keywords)
				retun=method(browser, driver, target, data)
				Keyword_execution_result_main=retun
				resultSet.append(Keyword_execution_result_main)
				createXLSReport(currentTestSuiteXLSPATH, TestStepsCount, Keyword_execution_result_main, dataset)
				
	printresults(currentTestSuiteXLSPATH, dataset, count, currentTestCase, resultSet, Priority, subdirectory)

def createXLSReport(currentTestSuiteXLSPATH, TestStepsCount, Keyword_execution_result_main, dataset):
	columnname="Result"+str(dataset-1)
	addColumn(currentTestSuiteXLSPATH, "TestSteps", columnname)
	addColumn(currentTestSuiteXLSPATH, "TestCases", columnname)
	addCellValue(currentTestSuiteXLSPATH, "TestSteps", TestStepsCount, columnname, Keyword_execution_result_main)

def printresults(currentTestSuiteXLSPATH, dataset, count, currentTestCase, resultSet, Priority, subdirectory):
	ResultsXLSPATH=subdirectory+'Results.xlsx'
	ResultsSheetXLS=load_workbook(ResultsXLSPATH).get_sheet_by_name('Status')
	columnname="Result"+str(dataset-1)
	if "FAIL" in str(resultSet):
		addCellValue(currentTestSuiteXLSPATH, "TestCases", count, columnname, "FAIL")
		addCellValue(currentTestSuiteXLSPATH, currentTestCase, dataset, "Result", "FAIL")
		FAIL=getCellValueBySheet(ResultsSheetXLS, 6, "Result")
		addCellValue(ResultsXLSPATH, "Status", 6, "Result", int(FAIL)+1)
		
		P1=getCellValueBySheet(ResultsSheetXLS, 19, "Result")
		P2=getCellValueBySheet(ResultsSheetXLS, 20, "Result")
		P3=getCellValueBySheet(ResultsSheetXLS, 21, "Result")
		P4=getCellValueBySheet(ResultsSheetXLS, 22, "Result")
		if Priority=="P1":
			addCellValue(ResultsXLSPATH, "Status", 19, "Result", int(P1)+1)
		elif Priority=="P2":
			addCellValue(ResultsXLSPATH, "Status", 20, "Result", int(P2)+1)
		elif Priority=="P3":
			addCellValue(ResultsXLSPATH, "Status", 21, "Result", int(P3)+1)
		elif Priority=="P4":
			addCellValue(ResultsXLSPATH, "Status", 22, "Result", int(P4)+1)
	else:
		addCellValue(currentTestSuiteXLSPATH, "TestCases", count, columnname, "PASS")
		addCellValue(currentTestSuiteXLSPATH, currentTestCase, dataset, "Result", "PASS")
		PASS=getCellValueBySheet(ResultsSheetXLS, 5, "Result")
		addCellValue(ResultsXLSPATH, "Status", 5, "Result", int(PASS)+1)
		
	#StartTime=getCellValueBySheet(ResultsSheetXLS, 5, "Results10")
	#ResultsXLSPATH=subdirectory+'Results.xlsx'
	#ResultsSheetXLS=load_workbook(ResultsXLSPATH).get_sheet_by_name('Status')
	global StartTime
	currentdate=StartTime.strftime('%d-%b-%Y')
	EndTime=datetime.datetime.now()
	
	addCellValue(ResultsXLSPATH, "Status", 4, "Results10", currentdate)
	addCellValue(ResultsXLSPATH, "Status", 5, "Results10", StartTime.strftime('%X'))
	#addCellValue(ResultsXLSPATH, "Status", 8, "Results10", browser)
	addCellValue(ResultsXLSPATH, "Status", 6, "Results10", EndTime.strftime('%X'))
	elapsed=EndTime-StartTime
	addCellValue(ResultsXLSPATH, "Status", 7, "Results10", elapsed)
	
def send_mail(AttachmentsDirPath):
    global logger
    logger.info('Gathering file and creating mail')
    msg=MIMEMultipart()
    msg['Subject']='Tesf file attached'
    msg['From']='techops@mtuity.com'
    recipients=['mohan.nimmala@mtuity.com', 'rithesh.karra@mtuity.com']
    msg['To']=", ".join(recipients)
    msg.attach(MIMEText( "Test mail"))
    for file in os.listdir(AttachmentsDirPath):
        filepath=os.path.join(AttachmentsDirPath,file)
        if not os.path.isfile(filepath):
           continue
        msgpart=MIMEBase('application',"vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        fileop=open(filepath,'rb')
        msgpart.set_payload(fileop.read())
        fileop.close()
        Encoders.encode_base64(msgpart)
        msgpart.add_header('Conent Disposition','attachment',filename=filepath)
        msg.attach(msgpart)
    logger.info('Connecting to SMTP  server')
    server=smtplib.SMTP('smtp.gmail.com:587')
    server.starttls()
    server.login('techops@mtuity.com','Paradigm@123.')
    server.sendmail(msg['From'],msg['To'],msg.as_string())
    logger.info('mail sent with attachments')
    server.quit()
		   		
# open a public URL, in this case, the webbrowser docs
def mainp():
    try:
    
        #t1=Thread(target=executefunctions,args=('FF', ''))
        t2=Thread(target=executefunctions,args=('Chrome', ''))
        #t3=Thread(target=executefunctions,args=('IE', ''))
        #t1.start()
        t2.start()
        #t3.start()
    except:
        print "Error: unable to start thread"

mainp()
