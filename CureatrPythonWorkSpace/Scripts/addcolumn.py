
#! /usr/bin/python
import os,sys
import shutil
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import  coordinate_from_string,column_index_from_string,get_column_letter


def addColumn(wbpath, columnname):
	print "add column function started"
	wb=load_workbook(wbpath)
	sh=wb.get_sheet_by_name("TestSteps")
	cl=sh.max_column
	print cl
	#sh.cell(row=1,column=cl-1).style.fill.start_color=Color('FF000000')
	sh.cell(row=1,column=cl+1).value=columnname
	#a.value=columnname
	#a.style.font.color='FF000000'
	wb.save(wbpath)
	print "add column function completed"
	
addColumn("/Users/macmini/Cureatr/CureatrPythonWorkSpace/OutPutFiles/2016-04-21/Chrome-Results18:56:31/Web_SIGNIN.xlsx", "Result1")