#! /usr/bin/python
import os,sys
import shutil
import openpyxl
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import  coordinate_from_string,column_index_from_string,get_column_letter

#src="/Users/macmini/Cureatr/CureatrPythonWorkSpace/InputFiles/Web/Web_SIGNIN.xlsx"
#dst="/Users/macmini/Cureatr/CureatrPythonWorkSpace/OutPutFiles/2016-04-19/"
#wb = load_workbook('/Users/macmini/Cureatr/CureatrPythonWorkSpace/InputFiles/Web/Web_SIGNIN.xlsx')
#sh = wb.worksheets[0]

def copyWorkBook(src, dst):
    shutil.copy(src, dst)

def getRowCount(sh):
    row_count= sh.max_row
    print row_count

def getColCount(sh):
    col_count=sh.max_column
    print col_count

def getCellData():
    for row in sh.iter_rows() :
        for cell in row:
            print cell.value
            
def Xlsx_Reader(FilePath, SheetName):
	wb=load_workbook(FilePath)
	sh=wb[SheetName]
	return sh
	
def issheetExist(FilePath, SheetName):
	 wb1 = load_workbook(FilePath)
	 listSheet=wb1.get_sheet_names()
	 for sname in range(0,len(listSheet)):
	 	if SheetName == listSheet[sname]:
	 		sheetExist="Yes"
	 		break
	 	else:
	 		sheetExist="No"
	 		
	 	
	 if sheetExist=="Yes":
	 	return True
	 else:
	 	return False
     		
     #if SheetName in str(listSheet):
     #		print "sheet exists"
     #		return True
     #else:
     #		#return False
     #		print "sheet  not exists"
     #		return False
     
    

def getCellValue(wb1, sh1, row, column):
     val=[]
     wb = load_workbook(wb1)
     sh2= wb[sh1]
     for srow in range(1, sh2.max_column+1):
         #for cell in srow:
             if sh2.cell(row=1,column=srow).value == column:
               val=sh2.cell(row=1, column=srow).col_idx
     return sh2.cell(row=row, column=val).value

def getCellValueBySheet(sh1, row, column):
     val=[]
     for srow in range(1, sh1.max_column+1):
         #for cell in srow:
             if sh1.cell(row=1,column=srow).value == column:
               val=sh1.cell(row=1, column=srow).col_idx
     return sh1.cell(row=row, column=val).value

def addColumn(wbpath, sh, column):
	wb=load_workbook(wbpath)
	sh=wb.get_sheet_by_name(sh)
	cl=sh.max_column
	for srow in range(1, sh.max_column+1):
         #for cell in srow:
         	if sh.cell(row=1,column=srow).value == column:
         		val=sh.cell(row=1, column=srow).col_idx
         		createsheet="No"
         		break
         	else:
         		createsheet="Yes"
         	
	if createsheet=="Yes":
		sh.cell(row=1,column=cl+1).value=column
		wb.save(wbpath)

def addCellValue(wbpath, sheet, row, column, columnvalue):
	val=[]
	wb=load_workbook(wbpath)
	sh=wb.get_sheet_by_name(sheet)
	for srow in range(1, sh.max_column+1):
         #for cell in srow:
             if sh.cell(row=1,column=srow).value == column:
               val=sh.cell(row=1, column=srow).col_idx
               sh.cell(row=row,column=val).value=columnvalue
               wb.save(wbpath)
               
